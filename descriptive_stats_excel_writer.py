# -*- coding: utf-8 -*-
"""
Created on Sun Mar 15 11:54:56 2026

@author: Pengzongjian
"""

# -*- coding: utf-8 -*-
"""
第6部分：结果导出到 Excel（字体优化版）
依赖：
- part1_data_loader.py
- part2_numeric_stats.py
- part3_categorical_stats.py
- part4_missing_stats.py

适用环境：Spyder / Python 3.x
功能：
1. 将统计结果导出到一个 Excel 文件
2. 自动设置表头样式、列宽、对齐方式
3. 中文使用宋体，英文和复杂文种使用 Times New Roman
4. 对混合文本单元格采用富文本分段写入
"""

import os
import pandas as pd

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont

from descriptive_stats_data_loader import (
    create_output_dirs,
    read_data_file,
    clean_data,
    identify_variable_types,
    summarize_data_info
)

from descriptive_stats_numeric_stats import generate_numeric_stats
from descriptive_stats_categorical_stats import generate_categorical_stats
from descriptive_stats_missing_stats import generate_missing_stats, generate_overall_missing_summary


# =========================
# 1. 通用辅助函数
# =========================

def create_notice_df(message):
    """
    当某类结果为空时，生成提示表，避免 Excel 工作表为空
    """
    return pd.DataFrame({"提示信息": [message]})


def is_cjk_char(ch):
    """
    判断字符是否属于中文/CJK及相关全角符号
    """
    code = ord(ch)
    return (
        0x3400 <= code <= 0x4DBF or   # CJK 扩展A
        0x4E00 <= code <= 0x9FFF or   # CJK 统一汉字
        0xF900 <= code <= 0xFAFF or   # CJK 兼容汉字
        0x3000 <= code <= 0x303F or   # CJK 符号和标点
        0xFF00 <= code <= 0xFFEF      # 全角符号
    )


def build_rich_text(text, font_size=11, bold=False):
    """
    将字符串拆分为：
    - 中文段：宋体
    - 英文/数字/复杂文种段：Times New Roman
    并返回 openpyxl 富文本对象
    """
    if text is None:
        return None

    text = str(text)
    if text == "":
        return ""

    runs = []
    current_text = text[0]
    current_is_cjk = is_cjk_char(text[0])

    for ch in text[1:]:
        ch_is_cjk = is_cjk_char(ch)
        if ch_is_cjk == current_is_cjk:
            current_text += ch
        else:
            font_name = "宋体" if current_is_cjk else "Times New Roman"
            runs.append(
                TextBlock(
                    InlineFont(rFont=font_name, sz=font_size, b=bold),
                    current_text
                )
            )
            current_text = ch
            current_is_cjk = ch_is_cjk

    # 最后一段
    font_name = "宋体" if current_is_cjk else "Times New Roman"
    runs.append(
        TextBlock(
            InlineFont(rFont=font_name, sz=font_size, b=bold),
            current_text
        )
    )

    return CellRichText(runs)


def estimate_display_width(text):
    """
    估算单元格显示宽度：
    中文按2个宽度估算，英文按1个宽度估算
    """
    if text is None:
        return 0

    text = str(text)
    width = 0
    for ch in text:
        width += 2 if is_cjk_char(ch) else 1
    return width


def auto_adjust_column_width(writer, sheet_name, df, extra_width=2, max_width=45):
    """
    自动调整 Excel 工作表列宽
    """
    worksheet = writer.sheets[sheet_name]

    for idx, col in enumerate(df.columns, start=1):
        values = [str(col)]
        if not df.empty:
            values += df[col].astype(str).tolist()

        max_len = max(estimate_display_width(v) for v in values) if values else 10
        adjusted_width = min(max_len + extra_width, max_width)

        column_letter = worksheet.cell(row=1, column=idx).column_letter
        worksheet.column_dimensions[column_letter].width = adjusted_width


# =========================
# 2. 工作表样式设置
# =========================

def apply_worksheet_style(ws):
    """
    统一设置工作表样式
    """
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    thin_gray = Side(style="thin", color="BFBFBF")

    # 冻结首行
    ws.freeze_panes = "A2"

    # 首行高度
    ws.row_dimensions[1].height = 24

    for row in ws.iter_rows():
        for cell in row:
            # 边框
            cell.border = Border(
                left=thin_gray,
                right=thin_gray,
                top=thin_gray,
                bottom=thin_gray
            )

            # 表头
            if cell.row == 1:
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                if cell.value is not None:
                    cell.value = build_rich_text(cell.value, font_size=11, bold=True)

            # 数据区
            else:
                # 对齐
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # 字体处理
                if isinstance(cell.value, str):
                    cell.value = build_rich_text(cell.value, font_size=11, bold=False)
                else:
                    # 纯数值/空值等，直接设 Times New Roman
                    cell.font = Font(name="Times New Roman", size=11)

    # 工作表默认网格线保留，可按需关闭：
    # ws.sheet_view.showGridLines = False


# =========================
# 3. Excel 导出
# =========================

def export_results_to_excel(
    output_excel_path,
    numeric_stats_df,
    categorical_stats_df,
    missing_stats_df,
    overall_missing_summary_df
):
    """
    将所有统计结果导出到一个 Excel 文件中，并设置字体格式
    """
    with pd.ExcelWriter(output_excel_path, engine="openpyxl") as writer:
        # 1. 数值型描述统计
        if numeric_stats_df is None or numeric_stats_df.empty:
            numeric_output_df = create_notice_df("未识别到数值型变量，未生成数值型描述统计结果。")
        else:
            numeric_output_df = numeric_stats_df.copy()

        numeric_output_df.to_excel(writer, sheet_name="数值型描述统计", index=False)
        auto_adjust_column_width(writer, "数值型描述统计", numeric_output_df)

        # 2. 分类型频数统计
        if categorical_stats_df is None or categorical_stats_df.empty:
            categorical_output_df = create_notice_df("未识别到分类型变量，未生成分类型频数统计结果。")
        else:
            categorical_output_df = categorical_stats_df.copy()

        categorical_output_df.to_excel(writer, sheet_name="分类型频数统计", index=False)
        auto_adjust_column_width(writer, "分类型频数统计", categorical_output_df)

        # 3. 每列缺失值统计
        if missing_stats_df is None or missing_stats_df.empty:
            missing_output_df = create_notice_df("未生成缺失值统计结果。")
        else:
            missing_output_df = missing_stats_df.copy()

        missing_output_df.to_excel(writer, sheet_name="缺失值统计", index=False)
        auto_adjust_column_width(writer, "缺失值统计", missing_output_df)

        # 4. 整体缺失汇总
        if overall_missing_summary_df is None or overall_missing_summary_df.empty:
            overall_output_df = create_notice_df("未生成整体缺失情况汇总结果。")
        else:
            overall_output_df = overall_missing_summary_df.copy()

        overall_output_df.to_excel(writer, sheet_name="整体缺失汇总", index=False)
        auto_adjust_column_width(writer, "整体缺失汇总", overall_output_df)

        # 5. 样式统一处理
        for sheet_name in writer.book.sheetnames:
            ws = writer.book[sheet_name]
            apply_worksheet_style(ws)

    print(f"Excel 结果文件已导出：{output_excel_path}")


# =========================
# 4. 主函数
# =========================

def main():
    """
    独立测试本模块
    """
    # ====== 这里改成你自己的路径 ======
    file_path = r"C:\Users\彭宗健\Desktop\软件设置\试验数据.xlsx"
    output_dir = r"C:\Users\彭宗健\Desktop\软件设置\输出结果\第六步输出结果"
    sheet_name = None

    try:
        # 1. 创建输出目录
        create_output_dirs(output_dir)

        # 2. 读取数据
        df, file_name = read_data_file(file_path, sheet_name=sheet_name)

        # 3. 数据预处理
        df = clean_data(df)

        # 4. 识别变量类型
        numeric_cols, categorical_cols = identify_variable_types(df)

        # 5. 输出数据基本信息
        summarize_data_info(df, numeric_cols, categorical_cols)

        # 6. 生成各类统计结果
        numeric_stats_df = generate_numeric_stats(df, numeric_cols, round_digits=4)
        categorical_stats_df = generate_categorical_stats(df, categorical_cols, round_digits=4)
        missing_stats_df = generate_missing_stats(
            df,
            numeric_cols=numeric_cols,
            categorical_cols=categorical_cols,
            round_digits=4
        )
        overall_missing_summary_df = generate_overall_missing_summary(df, round_digits=4)

        # 7. 设置输出 Excel 路径
        excel_output_dir = os.path.join(output_dir, "统计结果Excel")
        output_excel_path = os.path.join(excel_output_dir, f"{file_name}_描述统计结果.xlsx")

        # 8. 导出 Excel
        export_results_to_excel(
            output_excel_path=output_excel_path,
            numeric_stats_df=numeric_stats_df,
            categorical_stats_df=categorical_stats_df,
            missing_stats_df=missing_stats_df,
            overall_missing_summary_df=overall_missing_summary_df
        )

        print("\n第6部分（字体优化版）运行完成。")

    except Exception as e:
        print(f"程序运行出错：{e}")


if __name__ == "__main__":
    main()